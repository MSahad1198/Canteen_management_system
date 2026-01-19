import json
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Product, Order, OrderItem, Payment
from django.contrib.auth import authenticate, login, logout, login as auth_login, logout as auth_logout       # ✅ for authentication
from django.contrib import messages                                # ✅ for flash messages
from django.contrib.auth.forms import AuthenticationForm          # ✅ built-in login form
from django.shortcuts import redirect
from .models import Order
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Sum, Count
from django.contrib.admin.views.decorators import staff_member_required
from .models import Combo, Product
from datetime import timedelta, date
#exort to pdf import 
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from io import BytesIO
from django.db.models.functions import ExtractYear
import calendar
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import xlsxwriter




@login_required
def report(request):
    return render(request, 'canteen/report.html')

def is_admin(user):
    """Return True if user belongs to the 'Admin' group."""
    return user.groups.filter(name='Admin').exists()

def is_cashier(user):
    """Return True if user belongs to the 'Cashier' group."""
    return user.groups.filter(name='Cashier').exists()

@login_required
def index(request):
    is_admin = request.user.is_authenticated and request.user.groups.filter(name='Admin').exists()
    is_cashier = request.user.is_authenticated and request.user.groups.filter(name='Cashier').exists()
    return render(request, 'canteen/index.html', {
        'is_admin': is_admin,
        'is_cashier': is_cashier,
    })



@login_required
def pos_page(request):
    # POS frontend — will request products via AJAX
    return render(request, 'canteen/pos.html')

def api_products(request):
    qs = Product.objects.filter(in_stock=True).order_by('name')
    combos = Combo.objects.filter(show_in_pos=True)
    data = []

    for p in qs:
        data.append({
            'id': p.id,
            'name': p.name,
            'price': str(p.price),
            'stock_qty': p.stock_qty,
            'type': 'product',
        })
    for c in combos:
        data.append({
            'id': f"combo-{c.id}",
            'name': f"🍱 {c.name}",
            'price': str(c.final_price),
            'stock_qty': 'Combo',
            'type': 'combo',
        })
    return JsonResponse({'products': data})


@require_POST
@login_required
def api_create_order(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
        items = data.get('items', [])

        if not items:
            return HttpResponseBadRequest('No items provided')

        # ✅ Step 1: Validate stock first (before creating the order)
        insufficient = []  # list to collect items with low stock

        for it in items:
            pid = str(it['product_id'])
            qty = int(it.get('quantity', 1))

            if pid.startswith("combo-"):
                combo_id = pid.split("-")[1]
                combo = get_object_or_404(Combo, pk=combo_id)
                for prod in combo.items.all():
                    if prod.stock_qty < qty:
                        insufficient.append(f"{prod.name} (in {combo.name})")
            else:
                product = get_object_or_404(Product, pk=int(pid))
                if product.stock_qty < qty:
                    insufficient.append(product.name)

        # ❌ Stop if any product lacks stock
        if insufficient:
            return JsonResponse({
                "status": "error",
                "message": "Insufficient stock for: " + ", ".join(insufficient)
            }, status=400)

        # ✅ Step 2: Proceed to order creation
        total = Decimal('0')
        order = Order.objects.create(
            cashier=request.user,
            order_type=data.get('order_type', 'takeaway'),
            discount_amount=Decimal(str(data.get('discount_amount', '0'))),
            is_paid=False
        )

        # ✅ Step 3: Add items to order
        for it in items:
            pid = str(it['product_id'])
            qty = int(it.get('quantity', 1))

            if pid.startswith("combo-"):
                combo_id = pid.split("-")[1]
                combo = get_object_or_404(Combo, pk=combo_id)
                for prod in combo.items.all():
                    OrderItem.objects.create(
                        order=order,
                        product=prod,
                        quantity=qty,
                        unit_price=prod.price
                    )
                    total += prod.price * qty
            else:
                product = get_object_or_404(Product, pk=int(pid))
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=qty,
                    unit_price=product.price
                )
                total += product.price * qty

        total_after_discount = total - order.discount_amount
        order.total_amount = total_after_discount
        order.save()

        # ✅ Step 4: Process payment if provided
        payment_data = data.get('payment')
        if payment_data:
            paid_amount = Decimal(str(payment_data.get('paid_amount', '0')))
            change = Decimal('0')
            if payment_data.get('method') == 'cash':
                change = (paid_amount - total_after_discount) if paid_amount > total_after_discount else Decimal('0')

            Payment.objects.create(
                order=order,
                method=payment_data.get('method', 'cash'),
                paid_amount=paid_amount,
                change_given=change
            )

            order.is_paid = True
            order.save()  # stock updates trigger from signals

            # ===============================
# 🔔 Stock Alert Detection
# ===============================
            low_stock_products = []
            out_of_stock_products = []

            for item in order.items.all():
                product = item.product

            if product.stock_qty == 0:
                out_of_stock_products.append(product.name)

            elif product.stock_qty <= 5:
                low_stock_products.append({
                    "name": product.name,
                    "remaining": product.stock_qty
                })

        return JsonResponse({
             "status": "success",
              "order_id": order.id,
          # 🔔 alerts
                "low_stock": low_stock_products,
                "out_of_stock": out_of_stock_products
})


    except Exception as e:
        print("⚠️ Error creating order:", e)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def api_reprint_receipt(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    items = [{
        'product': it.product.name,
        'qty': it.quantity,
        'unit_price': str(it.unit_price),
        'line_total': str(it.line_total())
    } for it in order.items.all()]
    data = {
        'order_id': order.id,
        'created_at': order.created_at.isoformat(),
        'cashier': order.cashier.username if order.cashier else None,
        'items': items,
        'total': str(order.total_amount)
    }
    return JsonResponse(data)

@login_required
@user_passes_test(is_admin)
def daily_sales_report(request):
    # basic daily report for today
    today = timezone.localdate()
    orders = Order.objects.filter(created_at__date=today, is_paid=True, cancelled=False)
    total = sum(o.total_amount for o in orders)
    best_sellers = {}
    for o in orders:
        for it in o.items.all():
            best_sellers.setdefault(it.product.name, 0)
            best_sellers[it.product.name] += it.quantity
    # convert to sorted list
    best = sorted(best_sellers.items(), key=lambda x: -x[1])[:10]
    return JsonResponse({
        'date': str(today),
        'orders_count': orders.count(),
        'total_sales': str(total),
        'best_sellers': best
    })


def login(request):
    """
    Login page with role selection (Admin or Cashier).
    Admin → Django Admin panel
    Cashier → POS page
    """
    if request.user.is_authenticated:
        return redirect('canteen:index')

    if request.method == 'POST':
        role = request.POST.get('role')
        form = AuthenticationForm(request, data=request.POST)

        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)

            if user is not None:
                auth_login(request, user)

                # ✅ Add user to selected role group (optional)
                if role == 'admin':
                    group, _ = Group.objects.get_or_create(name='Admin')
                    user.groups.add(group)
                    return redirect('canteen:index')  # Django admin interface

                elif role == 'cashier':
                    group, _ = Group.objects.get_or_create(name='Cashier')
                    user.groups.add(group)
                    return redirect('canteen:cashier_dashboard')


                else:
                    return redirect('canteen:index')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid credentials.")
    else:
        form = AuthenticationForm()

    return render(request, 'canteen/login.html', {'form': form})


def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')


@login_required
def view_orders(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'canteen/view_orders.html', {'orders': orders})

def is_cashier(user):
    """Return True if user belongs to 'Cashier' group."""
    return user.groups.filter(name='Cashier').exists()

def is_admin(user):
    """Return True if user belongs to 'Admin' group."""
    return user.groups.filter(name='Admin').exists()

@login_required
def cashier_dashboard(request):
    """Dashboard for cashiers showing their sales summary and recent orders."""
    from .models import Order

    today = timezone.localdate()

    # Filter only this cashier's orders for today
    orders_today = Order.objects.filter(
        cashier=request.user,
        created_at__date=today
    )

    total_sales = orders_today.filter(is_paid=True).aggregate(total=Sum('total_amount'))['total'] or 0
    total_orders = orders_today.count()
    pending_orders = orders_today.filter(is_paid=False).count()

    # Show the last 10 orders
    recent_orders = orders_today.order_by('-created_at')[:10]

    context = {
        'total_sales': f"{total_sales:.2f}",
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'recent_orders': recent_orders,
    }
    return render(request, 'canteen/cashier_dashboard.html', context)

@user_passes_test(is_admin)
@login_required
def manage_combos(request):
    products = Product.objects.filter(in_stock=True)
    combos = Combo.objects.all().order_by('-created_at')
    return render(request, 'canteen/manage_combos.html', {
        'products': products,
        'combos': combos,
    })


@require_POST
@user_passes_test(is_admin)
@login_required
def api_create_combo(request):
    data = json.loads(request.body)
    combo = Combo.objects.create(
        name=data['name'],
        total_price=data['total_price'],
        discount_amount=data['discount_amount'],
        final_price=data['final_price']
    )
    combo.items.set(data['items'])
    combo.save()
    return JsonResponse({'message': '✅ Combo created successfully!'})


@require_POST
@login_required
@user_passes_test(is_admin)
def api_toggle_combo(request, combo_id):
    combo = get_object_or_404(Combo, pk=combo_id)
    combo.show_in_pos = not combo.show_in_pos
    combo.save()
    return JsonResponse({'status': combo.show_in_pos})

def reprint_receipt_view(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    return render(request, 'canteen/receipt.html', {'order': order})


#Report View
@login_required
def report(request):
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())  # Monday
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)

    # Daily totals
    daily_sales = (
        Order.objects.filter(created_at__date=today)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )

    # Weekly totals
    weekly_sales = (
        Order.objects.filter(created_at__date__gte=week_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )

    # Monthly totals
    monthly_sales = (
        Order.objects.filter(created_at__date__gte=month_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )

    # ✅ Yearly totals (NEW)
    yearly_sales = (
        Order.objects.filter(created_at__date__gte=year_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )

    # Best-selling items
    top_items = (
        OrderItem.objects.values('product__name')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:5]
    )

    context = {
        "daily_sales": daily_sales,
        "weekly_sales": weekly_sales,
        "monthly_sales": monthly_sales,
        "yearly_sales": yearly_sales,  # ✅ pass to template
        "top_items": top_items,
    }

    return render(request, "canteen/report.html", context)


@login_required
def export_report_pdf(request):
    # Prepare data
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    daily_sales = (
        Order.objects.filter(created_at__date=today)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    weekly_sales = (
        Order.objects.filter(created_at__date__gte=week_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    monthly_sales = (
        Order.objects.filter(created_at__date__gte=month_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    top_items = (
        OrderItem.objects.values('product__name')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:10]
    )

    # Create PDF
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setTitle("Sales Report")

    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, 810, "Canteen Sales Report")
    p.setFont("Helvetica", 12)
    p.drawString(50, 780, f"Date: {today}")
    p.drawString(50, 760, f"Daily Sales: ৳{daily_sales['total'] or 0}")
    p.drawString(50, 740, f"Weekly Sales: ৳{weekly_sales['total'] or 0}")
    p.drawString(50, 720, f"Monthly Sales: ৳{monthly_sales['total'] or 0}")

    y = 680
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Top Selling Items")
    y -= 20
    p.setFont("Helvetica", 11)

    for item in top_items:
        p.drawString(60, y, f"{item['product__name']} - {item['total_qty']} sold")
        y -= 18
        if y < 50:
            p.showPage()
            y = 800

    p.showPage()
    p.save()

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="canteen_report_{today}.pdf"'
    response.write(pdf)
    return response


@login_required
def export_report_excel(request):
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    daily_sales = (
        Order.objects.filter(created_at__date=today)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    weekly_sales = (
        Order.objects.filter(created_at__date__gte=week_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    monthly_sales = (
        Order.objects.filter(created_at__date__gte=month_start)
        .aggregate(total=Sum('total_amount'), count=Count('id'))
    )
    top_items = (
        OrderItem.objects.values('product__name')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:10]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws.append(["Canteen Sales Report"])
    ws.append(["Date", str(today)])
    ws.append([""])
    ws.append(["Daily Sales", daily_sales['total'] or 0])
    ws.append(["Weekly Sales", weekly_sales['total'] or 0])
    ws.append(["Monthly Sales", monthly_sales['total'] or 0])
    ws.append([""])
    ws.append(["Top Selling Items"])
    ws.append(["Item Name", "Quantity Sold"])

    for item in top_items:
        ws.append([item['product__name'], item['total_qty']])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="canteen_report_{today}.xlsx"'

    with BytesIO() as buffer:
        wb.save(buffer)
        response.write(buffer.getvalue())

    return response

@login_required
def cancel_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    # Only admin or the cashier who created the order can cancel
    if not (request.user.is_staff or order.cashier == request.user):
        return render(request, 'canteen/403.html')

    if request.method == "POST":
        reason = request.POST.get('cancel_reason')

        if not reason:
            messages.error(request, "Cancellation reason is required.")
            return redirect('canteen:cancel_order', order_id=order.id)

        order.cancelled = True
        order.cancel_reason = reason
        order.is_paid = False
        order.save()

        messages.success(request, f"Order #{order.id} cancelled successfully.")
        return redirect('canteen:view_orders')

    return render(request, 'canteen/cancel_order.html', {'order': order})

# daily sale page view

@login_required
def today_sales_detail(request):
    """
    View to display a detailed list of all sales that happened today.
    """
    today = timezone.localdate()
    
    # Filter for today's paid and non-cancelled orders
    orders = Order.objects.filter(
        created_at__date=today, 
        is_paid=True, 
        cancelled=False
    ).order_by('-created_at')
    
    # Calculate totals
    total_sales = sum(o.total_amount for o in orders)
    order_count = orders.count()
    
    context = {
        'today': today,
        'orders': orders,
        'total_sales': total_sales,
        'order_count': order_count,
    }
    return render(request, 'canteen/today_sales_detail.html', context)

# weekly sales detail view
@login_required
def weekly_sales_detail(request):
    """
    View to display sales for the last 7 days and top 3 trending products.
    """
    today = timezone.localdate()
    start_date = today - timedelta(days=7)  # Get date 7 days ago
    
    # Filter orders for the last 7 days
    orders = Order.objects.filter(
        created_at__date__range=[start_date, today],
        is_paid=True,
        cancelled=False
    ).order_by('-created_at')
    
    # Calculate Totals
    total_sales = sum(o.total_amount for o in orders)
    order_count = orders.count()

    # Calculate Trending Products (Top 3)
    product_sales = {}
    for order in orders:
        for item in order.items.all():
            name = item.product.name
            qty = item.quantity
            product_sales[name] = product_sales.get(name, 0) + qty
    
    # Sort by quantity (highest first) and take top 3
    # Result format: [('Burger', 50), ('Pizza', 30), ...]
    trending_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]

    context = {
        'start_date': start_date,
        'end_date': today,
        'orders': orders,
        'total_sales': total_sales,
        'order_count': order_count,
        'trending_products': trending_products,
    }
    return render(request, 'canteen/weekly_sales_detail.html', context)
  

@login_required
def yearly_sales_detail(request):
    """
    View to display sales breakdown for each month of the current year.
    Shows total sales and top 3 trending products per month.
    """
    current_year = timezone.now().year
    monthly_data = []

    # Loop through months 1 (January) to 12 (December)
    for month in range(1, 13):
        # Filter orders for this specific month and year
        orders = Order.objects.filter(
            created_at__year=current_year,
            created_at__month=month,
            is_paid=True,
            cancelled=False
        )
        
        # Calculate Total Sales for the month
        total_sales = sum(o.total_amount for o in orders)
        
        # Calculate Top 3 Products for the month
        product_sales = {}
        for order in orders:
            for item in order.items.all():
                name = item.product.name
                qty = item.quantity
                product_sales[name] = product_sales.get(name, 0) + qty
        
        # Sort and take top 3
        top_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
        
        monthly_data.append({
            'month_name': calendar.month_name[month],
            'total_sales': total_sales,
            'order_count': orders.count(),
            'top_products': top_products,
        })

    context = {
        'year': current_year,
        'monthly_data': monthly_data,
    }
    return render(request, 'canteen/yearly_sales_detail.html', context)


@login_required
def monthly_sales_detail(request):
    """
    Displays a calendar view for the current month.
    Includes sales data for previous/next month days if they fall in the same week.
    """
    today = timezone.localdate()
    year = today.year
    month = today.month

    # Setup Calendar (Start on Sunday = 6)
    cal = calendar.Calendar(firstweekday=6) 
    weeks = cal.monthdatescalendar(year, month)

    # 1. Optimize Query: Fetch orders for the entire visible range at once
    start_date = weeks[0][0]   # First day visible on calendar
    end_date = weeks[-1][-1]   # Last day visible on calendar

    all_orders = Order.objects.filter(
        created_at__date__range=[start_date, end_date],
        is_paid=True,
        cancelled=False
    ).prefetch_related('items__product')

    # 2. Group orders by date in a dictionary for fast lookup
    orders_by_date = {}
    for order in all_orders:
        d = order.created_at.date()
        if d not in orders_by_date:
            orders_by_date[d] = []
        orders_by_date[d].append(order)

    # 3. Build the Calendar Data Structure
    calendar_data = []
    for week in weeks:
        week_data = []
        for d in week:
            is_current_month = (d.month == month)
            day_orders = orders_by_date.get(d, [])
            
            # Calculate Totals
            total_sales = sum(o.total_amount for o in day_orders)
            
            # Calculate Trending (Top 3)
            product_sales = {}
            for o in day_orders:
                for item in o.items.all():
                    name = item.product.name
                    product_sales[name] = product_sales.get(name, 0) + item.quantity
            
            top_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]

            week_data.append({
                'date_obj': d,
                'day_number': d.day,
                'total_sales': total_sales,
                'top_products': top_products,
                'is_current_month': is_current_month,
                'is_today': (d == today)
            })
        calendar_data.append(week_data)

    context = {
        'month_name': calendar.month_name[month],
        'year': year,
        'calendar_data': calendar_data,
    }
    return render(request, 'canteen/monthly_sales_detail.html', context)


# In views.py - Add these imports at the top if not already present
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import xlsxwriter

# ============ TODAY'S SALES EXPORTS ============

@login_required
def export_today_sales_pdf(request):
    """Export today's sales to PDF"""
    today = timezone.localdate()
    orders = Order.objects.filter(
        created_at__date=today,
        is_paid=True,
        cancelled=False
    ).order_by('-created_at')
    
    total_sales = sum(o.total_amount for o in orders)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Today's Sales Report - {today.strftime('%B %d, %Y')}", title_style))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Summary
    summary_data = [
        ['Metric', 'Value'],
        ['Total Revenue', f'${total_sales}'],
        ['Total Orders', str(orders.count())],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Orders Table
    elements.append(Paragraph("Order Details", styles['Heading2']))
    elements.append(Spacer(1, 0.1 * inch))
    
    orders_data = [['Order ID', 'Time', 'Cashier', 'Items', 'Amount']]
    for order in orders:
        items_str = ', '.join([f"{it.quantity}x {it.product.name}" for it in order.items.all()])
        orders_data.append([
            str(order.id),
            order.created_at.strftime('%H:%M %p'),
            order.cashier.username if order.cashier else 'System',
            items_str[:50] + ('...' if len(items_str) > 50 else ''),
            f'${order.total_amount}'
        ])
    
    orders_table = Table(orders_data, colWidths=[1*inch, 1*inch, 1.2*inch, 1.5*inch, 1*inch])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(orders_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="todays_sales_{today}.pdf"'
    return response


@login_required
def export_today_sales_excel(request):
    """Export today's sales to Excel"""
    today = timezone.localdate()
    orders = Order.objects.filter(
        created_at__date=today,
        is_paid=True,
        cancelled=False
    ).order_by('-created_at')
    
    total_sales = sum(o.total_amount for o in orders)
    
    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Today's Sales")
    
    # Define styles
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#374151',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
    })
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
    })
    
    summary_format = workbook.add_format({
        'bold': True,
        'border': 1,
        'align': 'center',
    })
    
    data_format = workbook.add_format({
        'border': 1,
        'align': 'left',
    })
    
    currency_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '$#,##0.00',
    })
    
    # Title
    worksheet.merge_cells('A1:E1')
    worksheet.write('A1', f"Today's Sales Report - {today.strftime('%B %d, %Y')}", title_format)
    
    # Summary Section
    worksheet.write('A3', 'Metric', summary_format)
    worksheet.write('B3', 'Value', summary_format)
    worksheet.write('A4', 'Total Revenue', data_format)
    worksheet.write('B4', total_sales, currency_format)
    worksheet.write('A5', 'Total Orders', data_format)
    worksheet.write('B5', orders.count(), data_format)
    
    # Orders Table
    worksheet.write('A7', 'Order ID', header_format)
    worksheet.write('B7', 'Time', header_format)
    worksheet.write('C7', 'Cashier', header_format)
    worksheet.write('D7', 'Items', header_format)
    worksheet.write('E7', 'Amount', header_format)
    
    row = 7
    for order in orders:
        items_str = ', '.join([f"{it.quantity}x {it.product.name}" for it in order.items.all()])
        worksheet.write(row, 0, order.id, data_format)
        worksheet.write(row, 1, order.created_at.strftime('%H:%M %p'), data_format)
        worksheet.write(row, 2, order.cashier.username if order.cashier else 'System', data_format)
        worksheet.write(row, 3, items_str, data_format)
        worksheet.write(row, 4, float(order.total_amount), currency_format)
        row += 1
    
    # Set column widths
    worksheet.set_column('A:A', 12)
    worksheet.set_column('B:B', 12)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 40)
    worksheet.set_column('E:E', 12)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="todays_sales_{today}.xlsx"'
    return response


# ============ WEEKLY SALES EXPORTS ============

@login_required
def export_weekly_sales_pdf(request):
    """Export weekly sales to PDF"""
    today = timezone.localdate()
    start_date = today - timedelta(days=7)
    
    orders = Order.objects.filter(
        created_at__date__range=[start_date, today],
        is_paid=True,
        cancelled=False
    ).order_by('-created_at')
    
    total_sales = sum(o.total_amount for o in orders)
    
    # Calculate trending
    product_sales = {}
    for order in orders:
        for item in order.items.all():
            name = item.product.name
            product_sales[name] = product_sales.get(name, 0) + item.quantity
    trending_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Weekly Sales Report - {start_date.strftime('%b %d')} to {today.strftime('%b %d, %Y')}", title_style))
    elements.append(Spacer(1, 0.3 * inch))
    
    # Summary
    summary_data = [
        ['Metric', 'Value'],
        ['Weekly Revenue', f'${total_sales}'],
        ['Total Orders', str(orders.count())],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9333ea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Orders Table
    elements.append(Paragraph("Order Details", styles['Heading2']))
    elements.append(Spacer(1, 0.1 * inch))
    
    orders_data = [['Date', 'Order ID', 'Items', 'Amount']]
    for order in orders:
        items_str = ', '.join([f"{it.quantity}x {it.product.name}" for it in order.items.all()])
        orders_data.append([
            order.created_at.strftime('%m/%d %H:%M'),
            str(order.id),
            items_str[:50] + ('...' if len(items_str) > 50 else ''),
            f'${order.total_amount}'
        ])
    
    orders_table = Table(orders_data, colWidths=[1.2*inch, 1*inch, 2*inch, 1*inch])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9333ea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(orders_table)
    elements.append(Spacer(1, 0.3 * inch))
    
    # Trending Products
    elements.append(Paragraph("🔥 Top 3 Trending Products", styles['Heading2']))
    elements.append(Spacer(1, 0.1 * inch))
    
    trending_data = [['Product Name', 'Quantity Sold']]
    for idx, (name, qty) in enumerate(trending_products, 1):
        trending_data.append([f"{idx}. {name}", str(qty)])
    
    trending_table = Table(trending_data, colWidths=[3*inch, 2*inch])
    trending_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ea580c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(trending_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="weekly_sales_{today}.pdf"'
    return response


@login_required
def export_weekly_sales_excel(request):
    """Export weekly sales to Excel"""
    today = timezone.localdate()
    start_date = today - timedelta(days=7)
    
    orders = Order.objects.filter(
        created_at__date__range=[start_date, today],
        is_paid=True,
        cancelled=False
    ).order_by('-created_at')
    
    total_sales = sum(o.total_amount for o in orders)
    
    # Calculate trending
    product_sales = {}
    for order in orders:
        for item in order.items.all():
            name = item.product.name
            product_sales[name] = product_sales.get(name, 0) + item.quantity
    trending_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
    
    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Weekly Sales")
    
    # Styles
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#9333ea',
        'border': 1,
        'align': 'center',
    })
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
    })
    
    data_format = workbook.add_format({
        'border': 1,
        'align': 'left',
    })
    
    currency_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '$#,##0.00',
    })
    
    trending_header = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#ea580c',
        'border': 1,
        'align': 'center',
    })
    
    # Title
    worksheet.merge_cells('A1:D1')
    worksheet.write('A1', f"Weekly Sales Report - {start_date.strftime('%b %d')} to {today.strftime('%b %d, %Y')}", title_format)
    
    # Summary
    worksheet.write('A3', 'Metric', header_format)
    worksheet.write('B3', 'Value', header_format)
    worksheet.write('A4', 'Weekly Revenue', data_format)
    worksheet.write('B4', total_sales, currency_format)
    worksheet.write('A5', 'Total Orders', data_format)
    worksheet.write('B5', orders.count(), data_format)
    
    # Orders Table
    worksheet.write('A7', 'Date', header_format)
    worksheet.write('B7', 'Order ID', header_format)
    worksheet.write('C7', 'Items', header_format)
    worksheet.write('D7', 'Amount', header_format)
    
    row = 7
    for order in orders:
        items_str = ', '.join([f"{it.quantity}x {it.product.name}" for it in order.items.all()])
        worksheet.write(row, 0, order.created_at.strftime('%m/%d %H:%M'), data_format)
        worksheet.write(row, 1, order.id, data_format)
        worksheet.write(row, 2, items_str, data_format)
        worksheet.write(row, 3, float(order.total_amount), currency_format)
        row += 1
    
    # Trending Section
    row += 2
    worksheet.write(row, 0, 'Product Name', trending_header)
    worksheet.write(row, 1, 'Quantity Sold', trending_header)
    row += 1
    
    for idx, (name, qty) in enumerate(trending_products, 1):
        worksheet.write(row, 0, f"{idx}. {name}", data_format)
        worksheet.write(row, 1, qty, data_format)
        row += 1
    
    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:B', 12)
    worksheet.set_column('C:C', 40)
    worksheet.set_column('D:D', 12)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="weekly_sales_{today}.xlsx"'
    return response


# ============ MONTHLY SALES EXPORTS ============

@login_required
def export_monthly_sales_pdf(request):
    """Export monthly calendar sales to PDF"""
    today = timezone.localdate()
    year = today.year
    month = today.month
    
    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdatescalendar(year, month)
    
    start_date = weeks[0][0]
    end_date = weeks[-1][-1]
    
    all_orders = Order.objects.filter(
        created_at__date__range=[start_date, end_date],
        is_paid=True,
        cancelled=False
    ).prefetch_related('items__product')
    
    # Group orders by date
    orders_by_date = {}
    for order in all_orders:
        d = order.created_at.date()
        if d not in orders_by_date:
            orders_by_date[d] = []
        orders_by_date[d].append(order)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph(f"{calendar.month_name[month]} {year} - Calendar Report", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Calendar Table
    cal_data = [['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']]
    
    for week in weeks:
        week_row = []
        for d in week:
            day_orders = orders_by_date.get(d, [])
            total_sales = sum(o.total_amount for o in day_orders)
            
            # Build cell text
            cell_text = f"<b>{d.day}</b>"
            if total_sales > 0:
                cell_text += f"<br/>${total_sales}"
            
            # Calculate top product
            product_sales = {}
            for o in day_orders:
                for item in o.items.all():
                    product_sales[item.product.name] = product_sales.get(item.product.name, 0) + item.quantity
            
            if product_sales:
                top_product = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[0]
                cell_text += f"<br/><i>{top_product[0][:15]}</i>"
            
            week_row.append(cell_text)
        cal_data.append(week_row)
    
    cal_table = Table(cal_data, colWidths=[1.2*inch]*7)
    cal_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('MINROWHEIGHT', (0, 1), (-1, -1), 0.8*inch),
    ]))
    elements.append(cal_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="monthly_sales_{year}_{month:02d}.pdf"'
    return response


@login_required
def export_monthly_sales_excel(request):
    """Export monthly calendar sales to Excel"""
    today = timezone.localdate()
    year = today.year
    month = today.month
    
    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdatescalendar(year, month)
    
    start_date = weeks[0][0]
    end_date = weeks[-1][-1]
    
    all_orders = Order.objects.filter(
        created_at__date__range=[start_date, end_date],
        is_paid=True,
        cancelled=False
    ).prefetch_related('items__product')
    
    # Group orders by date
    orders_by_date = {}
    for order in all_orders:
        d = order.created_at.date()
        if d not in orders_by_date:
            orders_by_date[d] = []
        orders_by_date[d].append(order)
    
    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet(f"{calendar.month_name[month]}")
    
    # Styles
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter',
    })
    
    day_header = workbook.add_format({
        'bold': True,
        'bg_color': '#374151',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1,
    })
    
    day_cell = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'top',
        'text_wrap': True,
    })
    
    other_month_cell = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'top',
        'bg_color': '#d3d3d3',
        'text_wrap': True,
    })
    
    # Title
    worksheet.merge_cells('A1:G1')
    worksheet.write('A1', f"{calendar.month_name[month]} {year} - Calendar Report", title_format)
    
    # Day headers
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    for col, day in enumerate(days):
        worksheet.write(2, col, day, day_header)
    
    # Calendar cells
    row = 3
    for week in weeks:
        for d in week:
            is_current_month = (d.month == month)
            day_orders = orders_by_date.get(d, [])
            total_sales = sum(o.total_amount for o in day_orders)
            
            # Build cell content
            cell_content = f"{d.day}\n"
            if total_sales > 0:
                cell_content += f"${total_sales}\n"
            
            # Top product
            product_sales = {}
            for o in day_orders:
                for item in o.items.all():
                    product_sales[item.product.name] = product_sales.get(item.product.name, 0) + item.quantity
            
            if product_sales:
                top_product = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[0]
                cell_content += f"{top_product[0][:20]}"
            
            cell_format = day_cell if is_current_month else other_month_cell
            col = (weeks.index(week) * 7 + week.index(d)) % 7  # This needs adjustment
        
        # Simpler approach - write day by day
        for col, d in enumerate(week):
            is_current_month = (d.month == month)
            day_orders = orders_by_date.get(d, [])
            total_sales = sum(o.total_amount for o in day_orders)
            
            cell_content = f"{d.day}\n"
            if total_sales > 0:
                cell_content += f"${total_sales}\n"
            
            product_sales = {}
            for o in day_orders:
                for item in o.items.all():
                    product_sales[item.product.name] = product_sales.get(item.product.name, 0) + item.quantity
            
            if product_sales:
                top_product = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[0]
                cell_content += f"{top_product[0][:20]}"
            
            cell_format = day_cell if is_current_month else other_month_cell
            worksheet.write(row, col, cell_content, cell_format)
        
        row += 1
    
    # Set column widths
    for col in range(7):
        worksheet.set_column(col, col, 18)
    
    # Set row heights
    worksheet.set_row(0, 25)
    worksheet.set_row(2, 20)
    for r in range(3, row):
        worksheet.set_row(r, 60)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="monthly_sales_{year}_{month:02d}.xlsx"'
    return response


# ============ YEARLY SALES EXPORTS ============

@login_required
def export_yearly_sales_pdf(request):
    """Export yearly sales by month to PDF"""
    current_year = timezone.now().year
    
    monthly_data = []
    for month in range(1, 13):
        orders = Order.objects.filter(
            created_at__year=current_year,
            created_at__month=month,
            is_paid=True,
            cancelled=False
        )
        
        total_sales = sum(o.total_amount for o in orders)
        
        product_sales = {}
        for order in orders:
            for item in order.items.all():
                name = item.product.name
                product_sales[name] = product_sales.get(name, 0) + item.quantity
        
        top_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
        
        monthly_data.append({
            'month_name': calendar.month_name[month],
            'total_sales': total_sales,
            'order_count': orders.count(),
            'top_products': top_products,
        })
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Yearly Sales Report - {current_year}", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Monthly breakdown table
    monthly_table_data = [['Month', 'Revenue', 'Orders', 'Top Product', 'Qty']]
    
    for data in monthly_data:
        top_prod_name = data['top_products'][0][0] if data['top_products'] else 'N/A'
        top_prod_qty = data['top_products'][0][1] if data['top_products'] else 0
        
        monthly_table_data.append([
            data['month_name'],
            f"${data['total_sales']}",
            str(data['order_count']),
            top_prod_name[:20],
            str(top_prod_qty)
        ])
    
    monthly_table = Table(monthly_table_data, colWidths=[1.5*inch, 1.2*inch, 1*inch, 1.8*inch, 0.8*inch])
    monthly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(monthly_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="yearly_sales_{current_year}.pdf"'
    return response


@login_required
def export_yearly_sales_excel(request):
    """Export yearly sales by month to Excel"""
    current_year = timezone.now().year
    
    monthly_data = []
    for month in range(1, 13):
        orders = Order.objects.filter(
            created_at__year=current_year,
            created_at__month=month,
            is_paid=True,
            cancelled=False
        )
        
        total_sales = sum(o.total_amount for o in orders)
        
        product_sales = {}
        for order in orders:
            for item in order.items.all():
                name = item.product.name
                product_sales[name] = product_sales.get(name, 0) + item.quantity
        
        top_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
        
        monthly_data.append({
            'month_name': calendar.month_name[month],
            'total_sales': total_sales,
            'order_count': orders.count(),
            'top_products': top_products,
        })
    
    # Create workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Yearly Sales")
    
    # Styles
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
    })
    
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4f46e5',
        'font_color': 'white',
        'align': 'center',
        'border': 1,
    })
    
    data_format = workbook.add_format({
        'border': 1,
        'align': 'center',
    })
    
    currency_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'num_format': '$#,##0.00',
    })
    
    # Title
    worksheet.merge_cells('A1:E1')
    worksheet.write('A1', f"Yearly Sales Report - {current_year}", title_format)
    
    # Headers
    worksheet.write('A3', 'Month', header_format)
    worksheet.write('B3', 'Revenue', header_format)
    worksheet.write('C3', 'Orders', header_format)
    worksheet.write('D3', 'Top Product', header_format)
    worksheet.write('E3', 'Qty', header_format)
    
    # Data
    row = 3
    for data in monthly_data:
        top_prod_name = data['top_products'][0][0] if data['top_products'] else 'N/A'
        top_prod_qty = data['top_products'][0][1] if data['top_products'] else 0
        
        worksheet.write(row, 0, data['month_name'], data_format)
        worksheet.write(row, 1, float(data['total_sales']), currency_format)
        worksheet.write(row, 2, data['order_count'], data_format)
        worksheet.write(row, 3, top_prod_name, data_format)
        worksheet.write(row, 4, top_prod_qty, data_format)
        row += 1
    
    worksheet.set_column('A:A', 15)
    worksheet.set_column('B:B', 15)
    worksheet.set_column('C:C', 12)
    worksheet.set_column('D:D', 25)
    worksheet.set_column('E:E', 10)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="yearly_sales_{current_year}.xlsx"'
    return response
